# -*- coding: utf-8 -*-


# -*- coding: utf-8 -*-
"""
Created on Sat Mar  6 19:59:27 2021

@author: CHENLF
"""
# -*- coding: utf-8 -*-
"""
Created on Sat Mar  6 19:59:27 2021

@author: CHENLF
"""
from tkinter import *
from tkinter.filedialog import  askdirectory, askopenfilenames, askopenfilename
from tkinter.messagebox import showinfo
import datetime, os
import os, sys
from docxtpl import DocxTemplate
import queue, time, datetime
from docx import Document
from xltpl.writerx import BookWriter
from xlrd import xldate_as_tuple
import time, wmi
import locale
locale.setlocale(locale.LC_ALL,'en') 
locale.setlocale(locale.LC_CTYPE,'chinese')
import win32api
import xlrd
import openpyxl
import win32com
from win32com.client import Dispatch, constants

def creat_ydj(file, data, path):
    name = os.path.split(file)[-1]
    save_path = os.path.join(path, data + name)
    workbook = openpyxl.load_workbook(file, keep_vba=True)
    sheet = workbook.get_sheet_by_name('产品信息详表')
    sheet.cell(row = 18, column = 3).value = data  
    workbook.save(save_path)

if not os.path.exists(r'C:\Files'):
    os.mkdir(r'C:\Files')

facebg = 'DarkGray'  # Honeydew  \PaleGoldenrod \PowderBlue \Silver                     #窗口背景色，其他背景色见：blog.csdn.net/chl0000/article/details/7657887
menubg = 'DarkGray' ##CornflowerBlue\GhostWhite\

def get_data(element_file, sheet_index = 0, nor = 0):
	## 设置文件读取函数
    element_file = element_file
    import xlrd
    data = xlrd.open_workbook(element_file)
    table = data.sheet_by_index(sheet_index)
    if nor == 0:
        nor = table.nrows - 1
    else:
        nor = nor
    ncols = table.ncols
    dict = {}
    for col in range(ncols):
        if table.cell_value(0, col):
            if table.cell(nor, col).ctype == 0:
            	pass
                #dict[table.cell_value(0, col)] = "{{" + table.cell_value(0, col)  + "}}"
            elif table.cell(nor, col).ctype == 3:
                date = datetime.datetime(*xldate_as_tuple(table.cell_value(nor, col), 0))
                datestr = str(date.year) + '年' + str(date.month) + '月'+ str(date.day) + '日'
                dict[table.cell_value(0, col)] = datestr 
            elif type(table.cell_value(nor, col)) == float or type(table.cell_value(nor, col)) == int:
                if table.cell_value(nor, col) == 2020 or table.cell_value(nor, col) == 2021 or table.cell_value(nor, col) == 2019 or table.cell_value(nor, col) == 2022:
                    try:
                        dict[table.cell_value(0, col)] = int(table.cell_value(nor, col))
                    except:
                        dict[table.cell_value(0, col)] = table.cell_value(nor, col)
                elif table.cell_value(nor, col) > 3:
                    try:
                        dict[table.cell_value(0, col)] = format(float(table.cell_value(nor, col)),',.2f')
                    except:
                        dict[table.cell_value(0, col)] = table.cell_value(nor, col)  
                else:
                    try:
                        dict[table.cell_value(0, col)] = format(float(table.cell_value(nor, col)),',.4f')
                    except:
                        dict[table.cell_value(0, col)] = table.cell_value(nor, col)
            else:
                dict[table.cell_value(0, col)] = table.cell_value(nor, col)
    return dict


def clean(data):
    #print(data)
    keys = data.keys()
    keypop = []
    for key in keys:
        if  type(data[key])  == str:
            if len(data[key]) <=10:
                if re.search(r"(\d{4}-\d{1,2}-\d{1,2})",data[key]):
                    data[key] = datetime.datetime.strftime(datetime.datetime.strptime(data[key].replace(' ',''),'%Y-%m-%d'),'%Y年%m月%d日')
        #elif data[key] == None:
           # data = data.pop(key)
    data2 = {}
    for o in data:
        if not data[o] == None:
            data2[o] = data[o]
    data2['today'] = datetime.datetime.strftime(datetime.datetime.today(),'%Y年%m月%d日')
    #print(data2)
    return data2




LOG_LINE_NUM = 0
global path_
path_ = r'C:\Files'

class MY_GUI():
    def __init__(self,init_window_name):
        self.init_window_name = init_window_name

    #设置窗口
    def set_init_window(self):
        width = 600
        height = 680
        self.init_window_name.iconbitmap('.\\rocket3.ico')
        self.init_window_name.title("MapFiles-Excle")           #窗口名
        self.init_window_name.resizable(width=False, height=False)
        self.screenwidth = self.init_window_name.winfo_screenwidth()
        self.screenheight = self.init_window_name.winfo_screenheight()
        self.alignstr = '%dx%d+%d+%d' % (width, height, (self.screenwidth-width)/2, (self.screenheight-height)/2)
        self.init_window_name.geometry(self.alignstr)

        self.init_window_name["bg"] = facebg           # Honeydew  \PaleGoldenrod \PowderBlue                      #窗口背景色，其他背景色见：blog.csdn.net/chl0000/article/details/7657887
        self.init_window_name.attributes("-alpha",1)                          #虚化，值越小虚化程度越高
        # 变量
        self.file_temp = StringVar()
        self.data = StringVar()
        self.path_save = StringVar()
        self.element = StringVar()
        #标签
        self.TableName = Label(self.init_window_name, text="Table：", bg = facebg,fg="Black",font=("Times New Roman", 11)).place(x=30, y= 290,  height=30)
        self.Data_Column = Label(self.init_window_name, text="Row：",bg = facebg, fg="Black",font=("Times New Roman", 11)).place(x=225, y= 290,  height=30)
        self.File_Name = Label(self.init_window_name, text="Names：",bg = facebg, fg="Black",font=("Times New Roman", 11)).place(x=410, y= 290,  height=30)
        #self.rizhi = Label(self.init_window_name, text="日      志",bg = 'PowderBlue',fg="Black",font=("楷体", 13)).place(x=250, y= 355 ,  height=30)
        Label(self.init_window_name,text = "Copyright © 2020-2021 Chen Liangfang, All rights reserved.", font=("Times New Roman", 11), fg="Black", bg = facebg).place(x=105, y= 640)

        #按钮
        self.Path_Button = Button(self.init_window_name, text = "保存路径", bg = menubg, fg = 'Black', width=10, height=1, command = self.selectPath, bd = 3).place(x=485, y= 240)
        self.Temp_Button = Button(self.init_window_name, text = "模板选择", bg = menubg, fg = 'Black',width=10, height=1, command = self.selectFile, bd = 3).place(x=485, y= 90)
        self.Data_Button = Button(self.init_window_name, text = "数据源", bg = menubg, fg = 'Black',width=10, height=1, command = self.selectExcle, bd = 3).place(x=485, y= 190)

        self.Files_Button = Button(self.init_window_name, text = "生成文件", bg = menubg, fg = 'Black', width=20, height=2,command = self.mappfile, activebackground = 'orange', bd = 3).place(x=60, y= 340)
        self.Files_Button = Button(self.init_window_name, text = "生成PDF", bg = menubg, fg = 'Black', width=20, height=2,command = self.mapfile2pdf, activebackground = 'orange', bd = 3).place(x=380, y= 340)
        
        #文本框
        self.TableName_Entry = Entry(self.init_window_name, width=12,  bd = 3)
        self.Data_Column_Entry = Entry(self.init_window_name, width=12,  bd = 3)
        self.File_Name_Entry = Entry(self.init_window_name, width=12,  bd = 3)
        self.TableName_Entry.place(x=85, y= 290,  height=30)
        self.Data_Column_Entry.place(x=275, y= 290,  height=30)
        self.File_Name_Entry.place(x=475, y= 290,  height=30)
        self.TableName_Entry.insert(0,'Table')
        self.Data_Column_Entry.insert(0,'Row')
        self.File_Name_Entry.insert(0,'{{项目编号}}-@-{{项目简称}}-@')
        self.Temp_Date = Listbox(self.init_window_name, listvariable  = self.file_temp,width=60,height=8, bd = 3).place(x=30, y= 20)
        self.Save_path = Entry(self.init_window_name, textvariable = self.path_save,width=60, bd = 3)
        self.Save_path.place(x=30, y= 240, height=30)
        self.Data = Listbox(self.init_window_name, listvariable  = self.data,width=60, bd = 3)
        self.Data.place(x=30, y= 190, height=30)
        self.Save_path.insert(0,r'C:\Files')
        path_ = self.Save_path.get()
        self.log_data_Text = Text(self.init_window_name, width=75, height=9, bd = 3)
        self.log_data_Text.place(x=30, y= 410, height=220)
        
    # 功能函数
    def selectPath(self):
        global path_
        path_ = askdirectory(title='Set Save Path')
        self.path_save.set(path_)
    
    def selectFile(self):
        global file_
        file_ = askopenfilenames(title='Select Word Files', filetypes=[('File', '*.docx'), ('File', '*.xlsm'), ('File', '*.xlsx')], initialdir=(os.path.expanduser('H:/')))
        self.file_temp.set(file_)
        
    def selectExcle(self):
        global excel_
        excel_ = askopenfilenames(title='Select Excel Files', filetypes=[ ('File', '*.xls'), ('File', '*.xlsx')], initialdir=(os.path.expanduser('H:/')))
        excel_ = excel_[0]
        #print(excel_)
        self.data.set(excel_)
    def get_data_from_excel(self, sheet_index = 0, nor = 0):
        data = xlrd.open_workbook(excel_)
        table = data.sheet_by_index(int(sheet_index)-1)
        if nor == 0:
            nor = table.nrows - 1
        else:
            nor = nor
        ncols = table.ncols
        dicts = {}
        for col in range(ncols):
            if table.cell_value(0, col):
                if table.cell(nor, col).ctype == 0:
                    pass
                	#dicts[table.cell_value(0, col)] = "{{" + table.cell_value(0, col)  + "}}"
                elif table.cell(nor, col).ctype == 3:
                    date = datetime.datetime(*xldate_as_tuple(table.cell_value(nor, col), 0))
                    datestr = str(date.year) + '年' + str(date.month) + '月'+ str(date.day) + '日'
                    dicts[table.cell_value(0, col)] = datestr 
                elif type(table.cell_value(nor, col)) == float or type(table.cell_value(nor, col)) == int:
                    if table.cell_value(nor, col) == 2020 or table.cell_value(nor, col) == 2021 or table.cell_value(nor, col) == 2019 or table.cell_value(nor, col) == 2022:
                        try:
                            dicts[table.cell_value(0, col)] = int(table.cell_value(nor, col))
                        except:
                            dicts[table.cell_value(0, col)] = table.cell_value(nor, col)
                    elif table.cell_value(nor, col) > 3:
                        try:
                            dicts[table.cell_value(0, col)] = format(float(table.cell_value(nor, col)),',.2f')
                        except:
                            dicts[table.cell_value(0, col)] = table.cell_value(nor, col)  
                    else:
                        try:
                            dicts[table.cell_value(0, col)] = format(float(table.cell_value(nor, col)),',.4f')
                        except:
                            dicts[table.cell_value(0, col)] = table.cell_value(nor, col)
                else:
                    dicts[table.cell_value(0, col)] = table.cell_value(nor, col)
        return dicts
        
    ''' 
    def get_data_from_setable(self, Tabelname):
        seatable = SeaTableAPI(api_token, 'https://cloud.seatable.cn')
        seatable.auth()
        rows = seatable.list_rows(Tabelname)
        return rows
    '''
    ## 生成文件函数
    def produce_word(self, context, path_of_MB,path_of_HT, renames):
        files = path_of_MB
        context = clean(context)
        if renames == '{{Names}}':
            renames = '{{项目简称}}'
        for file in files:
            if file.endswith('docx'):
                try:
                    tpl = DocxTemplate(file)
                    try:
                        tpl.render(context)
                    except Exception as e:
                        self.write_log_to_Text("INFO: %s faild" % (renames + os.path.splitext(os.path.basename(file))[0].replace('-模板','').replace('模板','')))
                        self.write_log_to_Text("错误原因：%s" %e)
                        showinfo(title = "提示",
                            message = "模板可能存在的问题：{0}".format(str(e)))
                        raise e
                    try:
                        tpl.save(os.path.join(path_of_HT,  renames + os.path.splitext(os.path.basename(file))[0].replace('-模板','').replace('模板','')   + os.path.splitext(os.path.basename(file))[1]))
                        self.write_log_to_Text("INFO: %s success" % (renames + os.path.splitext(os.path.basename(file))[0].replace('-模板','').replace('模板','')))
                    except Exception as e:
                        self.write_log_to_Text("INFO: %s" %e)
                        showinfo(title = "提示",
                            message = "文件保存文件时报错！")
                        self.write_log_to_Text("INFO: %s" %e)
                        raise e
                except Exception as e:
                    self.write_log_to_Text("INFO: %s" %e)
                    showinfo(title = "提示",
                            message = "模板可能存在问题！请检查{{}}的全角半角格式！")
                    raise e
            elif file.endswith('xlsx'):
                try:
                    tpl = BookWriter(file)
                    try:
                        tpl.render_sheet(context, 0, 0)
                    except Exception as e:
                        self.write_log_to_Text("INFO: %s faild" % (renames + os.path.splitext(os.path.basename(file))[0].replace('-模板','').replace('模板','')))
                        self.write_log_to_Text("INFO: %s" %e)
                        showinfo(title = "提示",
                            message = "模板可能存在的问题：{0}".format(str(e)))
                        raise e
                    try:
                        tpl.save(os.path.join(path_of_HT,renames + os.path.splitext(os.path.basename(file))[0].replace('-模板','').replace('模板','')   + os.path.splitext(os.path.basename(file))[1]))
                        self.write_log_to_Text("INFO: %s success" % (renames + os.path.splitext(os.path.basename(file))[0].replace('-模板','').replace('模板','')))
                    except Exception as e:
                        self.write_log_to_Text("INFO: %s" %e)
                        showinfo(title = "提示",
                            message = "文件保存文件时报错！")
                        raise e
                except Exception as e:
                    self.write_log_to_Text("INFO: %s" %e)
                    showinfo(title = "提示",
                            message = "模板可能存在问题！请检查{{}}的全角半角格式！")
                    raise e
            elif file.endswith('xlsm'):
                try:
                    name = os.path.split(file)[-1]
                    save_path = os.path.join(path_of_HT,renames + os.path.splitext(os.path.basename(file))[0].replace('-模板','').replace('模板','')   + os.path.splitext(os.path.basename(file))[1])
                    workbook = openpyxl.load_workbook(file, keep_vba=True)
                    sheet = workbook['产品信息详表']
                    #print(context)
                    #print(context['项目全称'])
                    try:
                        sheet.cell(row = 18, column = 3).value = context['项目全称']
                    except:
                        try:
                            sheet.cell(row = 18, column = 3).value = context['计划全称']
                        except:
                            try:
                                sheet.cell(row = 18, column = 3).value = '华润信托·%s集合资金信托计划' % context['计划简称']
                            except:
                                sheet.cell(row = 18, column = 3).value = '华润信托·%s集合资金信托计划' % context['项目简称']

                    workbook.save(save_path)
                    self.write_log_to_Text("生成%s成功" %save_path)

                    try:
                        xlApp = win32com.client.DispatchEx("Excel.Application")
                        xlApp.Visible = False
                        xlApp.DisplayAlerts = 0
                        xlBook = xlApp.Workbooks.Open(save_path,False)
                        xlBook.Application.Run("Butn_WriteJson_Factor_Click")#宏
                        xlBook.Close(True)                        
                        self.write_log_to_Text("导出%sjson数据成功" %save_path)

                    except:
                        self.write_log_to_Text("导出%sjson数据失败" %save_path)

                        
                except Exception as e:
                    self.write_log_to_Text("INFO: %s" %e)
                    showinfo(title = "提示",
                            message = "请保证项目全称存在且列名为：项目全称or计划全称")
                    raise e
                               
            else:
                self.write_log_to_Text("INFO: %s" % "模板格式有问题")
                showinfo(title = "提示",
                            message = "模板格式有问题") 
    def cleanword2pdf(self,file):
        # 获取清洁版PDF
        from datetime import datetime
        from win32com.client import Dispatch, constants
        w = Dispatch('Word.Application')
        w.Visible = 0
        w.DisplayAlerts = 0
        file = file.replace('/','\\')
        worddoc = w.Documents.Open(file)
        filename = os.path.splitext(os.path.basename(file))[0]
        docx = os.path.splitext(os.path.basename(file))[1]
        newdoc = filename + '-clean' + docx
        newpdf = filename + '.pdf'
        worddoc.ActiveWindow.Selection.Range.Revisions.AcceptAll()

        path = os.path.dirname(file).replace('/','\\')
        try:
            worddoc.DeleteAllComments()
        except:
            pass
        worddoc.ActiveWindow.Selection.WholeStory()  ## 实现全部选择
        worddoc.ActiveWindow.Selection.Range.HighlightColorIndex = 0 ## 设置背景色
        worddoc.ActiveWindow.Selection.Range.Font.Color = 0  ## 设置为黑色
        worddoc.AcceptAllRevisions()
        worddoc.TrackRevisions = False
        worddoc.AcceptAllRevisions()
        worddoc.AcceptAllRevisions()
        try:
            worddoc.SaveAs(os.path.join(path, newpdf),FileFormat= 17)
        except:
            os.remove(os.path.join(path, newpdf))
            worddoc.SaveAs(os.path.join(path, newpdf),FileFormat= 17)
        worddoc.Close() ## 原文档保留
        w.Quit()
    def produce_pdf(self, context, path_of_MB,path_of_HT, renames):
        files = path_of_MB
        context = clean(context)
        if renames == '{{Names}}':
            renames = '{{项目简称}}'
        for file in files:
            if file.endswith('docx'):
                try:
                    tpl = DocxTemplate(file)
                    try:
                        tpl.render(context)
                    except Exception as e:
                        self.write_log_to_Text("INFO: %s faild" % (renames + os.path.splitext(os.path.basename(file))[0].replace('-模板','').replace('模板','')))
                        self.write_log_to_Text("INFO: %s" %e)
                        showinfo(title = "提示",
                            message = "模板可能存在的问题：{0}".format(str(e)))
                        raise e
                    try:
                        tpl.save(os.path.join(path_of_HT,  renames + os.path.splitext(os.path.basename(file))[0].replace('-模板','').replace('模板','')   + os.path.splitext(os.path.basename(file))[1]))
                        self.cleanword2pdf(os.path.join(path_of_HT, renames + os.path.splitext(os.path.basename(file))[0].replace('-模板','').replace('模板','')   + os.path.splitext(os.path.basename(file))[1]))
                        self.write_log_to_Text("INFO: %s success" % (renames + os.path.splitext(os.path.basename(file))[0].replace('-模板','').replace('模板','')))
                    except Exception as e:
                        self.write_log_to_Text("INFO: %s" %e)
                        showinfo(title = "提示",
                            message = "请确认保存路径已经选择！")
                        raise e
                except Exception as e:
                    self.write_log_to_Text("INFO: %s" %e)
                    showinfo(title = "提示",
                            message = "请确认模板为docx格式！")
                    raise e
            elif file.endswith('xlsx'):
                try:
                    tpl = BookWriter(file)
                    try:
                        tpl.render_sheet(context, 0, 0)
                    except Exception as e:
                        self.write_log_to_Text("INFO: %s faild" % (renames + os.path.splitext(os.path.basename(file))[0].replace('-模板','').replace('模板','')))
                        self.write_log_to_Text("INFO: %s" %e)
                        showinfo(title = "提示",
                            message = "模板可能存在的问题：{0}".format(str(e)))
                        raise e
                    try:
                        tpl.save(os.path.join(path_of_HT,renames + os.path.splitext(os.path.basename(file))[0].replace('-模板','').replace('模板','')   + os.path.splitext(os.path.basename(file))[1]))
                        self.write_log_to_Text("INFO: %s success" % (renames + os.path.splitext(os.path.basename(file))[0].replace('-模板','').replace('模板','')))
                    except Exception as e:
                        self.write_log_to_Text("INFO: %s" %e)
                        showinfo(title = "提示",
                            message = "请确认保存路径已经选择！")
                        raise e
                except Exception as e:
                    self.write_log_to_Text("INFO: %s" %e)
                    showinfo(title = "提示",
                            message = "请确认模板为excle格式！")
                    raise e
            elif file.endswith('xlsm'):
                try:
                    name = os.path.split(file)[-1]
                    save_path = os.path.join(path_of_HT,renames + os.path.splitext(os.path.basename(file))[0].replace('-模板','').replace('模板','')   + os.path.splitext(os.path.basename(file))[1])
                    workbook = openpyxl.load_workbook(file, keep_vba=True)
                    sheet = workbook['产品信息详表']
                    #print(context)
                    #print(context['项目全称'])
                    try:
                        sheet.cell(row = 18, column = 3).value = context['项目全称']
                    except:
                        try:
                            sheet.cell(row = 18, column = 3).value = context['计划全称']
                        except:
                            try:
                                sheet.cell(row = 18, column = 3).value = '华润信托·%s集合资金信托计划' % context['计划简称']
                            except:
                                sheet.cell(row = 18, column = 3).value = '华润信托·%s集合资金信托计划' % context['项目简称']

                    workbook.save(save_path)
                    self.write_log_to_Text("生成%s成功" %save_path)

                    try:
                        xlApp = win32com.client.DispatchEx("Excel.Application")
                        xlApp.Visible = False
                        xlApp.DisplayAlerts = 0
                        xlBook = xlApp.Workbooks.Open(save_path,False)
                        xlBook.Application.Run("Butn_WriteJson_Factor_Click")#宏
                        xlBook.Close(True)                        
                        self.write_log_to_Text("导出%sjson数据成功" %save_path)

                    except:
                        self.write_log_to_Text("导出%sjson数据失败" %save_path)       
                except Exception as e:
                    self.write_log_to_Text("INFO: %s" %e)
                    showinfo(title = "提示",
                            message = "请保证项目全称存在且列名为：项目全称or计划全称")
                    raise e
                              
            else:
                self.write_log_to_Text("INFO: %s" %e)
                showinfo(title = "提示",
                            message = "请选择word格式和excel格式的文件！") 
    def mappfile(self):
        if datetime.datetime.today().year > 2022:
            showinfo(title = "提示",
                message = "使用期限已过，请联系作者陈良方！")
        else:
            if self.Data_Column_Entry.get() == 'Row':
                showinfo(title = "提示",
                                message = "Row框请输入数字！")
            else:
                try:
                    row_index = int(self.Data_Column_Entry.get()) - 1
                    row_index = [row_index]
                except Exception as e:
                    try:
                        row_index = self.Data_Column_Entry.get().replace(' ', '').replace('，', ',').split(',')
                        ints = lambda x:int(x) -1
                        row_index = list(map(ints, row_index))
                    except Exception as e:
                        try:
                            row_index = self.Data_Column_Entry.get().replace(' ', '').split('-')
                            row_index = range(int(row_index[0])-1, int(row_index[1]))
                        except Exception as e:
                            self.write_log_to_Text("INFO: %s" %e)
                            showinfo(title = "提示",
                                message = "Row框请输入数字,如果多行以半角,隔开！")
                            raise e
                for i in row_index:
                    try:
                        context = self.get_data_from_excel(self.TableName_Entry.get(),i)
                    except Exception as e:
                        self.write_log_to_Text("INFO: %s" %e)
                        showinfo(title = "提示",
                            message = "获取数据源失败，检查excel表格！")
                        raise e
                    try:
                        renames = str(self.File_Name_Entry.get())
                        renames =   renames.replace(' ', '').split('-')
                        Names = ''
                        for rename in renames:             
                            if '{{' in rename and '}}' in rename :
                                rename = rename.replace("{{",'').replace("}}",'')
                                if rename in context.keys():
                                    try:
                                        rename = context[rename]
                                    except Exception as e:
                                        self.write_log_to_Text("INFO: %s" %e)
                                        showinfo(title = "提示",
                                            message = "Name框请确定变量名是否正确")
                                else:
                                    #self.write_log_to_Text("INFO: %s" %e)
                                    showinfo(title = "提示",
                                            message = "命名字段[%s]不在Table中" %rename)
                            Names = Names + rename
                    except Exception as e:
                        self.write_log_to_Text("INFO: %s" %e)
                        showinfo(title = "提示",
                                    message = "Name框请输入文本")
                        raise e
                    try:
                        self.produce_word(context, file_, path_, Names)    
                    except Exception as e:
                        self.write_log_to_Text("INFO: %s" %e)
                        showinfo(title = "提示",
                                message = "请确认模板文件和保存路径已经选择！")
                        raise e       
                showinfo(title = "提示",
                          message = "文件已经生成！请在{0}查看".format(path_))
                os.startfile(path_)
        pass
    def mapfile2pdf(self):
        if datetime.datetime.today().year > 2022:
            showinfo(title = "提示",
                message = "使用期限已过，请联系作者陈良方！")
        else:
            if self.Data_Column_Entry.get() == 'Row':
                showinfo(title = "提示",
                                message = "Row框请输入数字,如果多行以半角,隔开！")
            else:
                try:
                    row_index = int(self.Data_Column_Entry.get()) - 1
                    row_index = [row_index]
                except Exception as e:
                    try:
                        row_index = self.Data_Column_Entry.get().replace(' ', '').replace('，', ',').split(',')
                        ints = lambda x:int(x) -1
                        row_index = list(map(ints, row_index))
                    except Exception as e:
                        try:
                            row_index = self.Data_Column_Entry.get().replace(' ', '').split('-')
                            row_index = range(int(row_index[0])-1, int(row_index[1]))
                        except Exception as e:
                            self.write_log_to_Text("INFO: %s" %e)
                            showinfo(title = "提示",
                                message = "Row框请输入数字,如果多行以半角,隔开！")
                            raise e
                '''
                try:
                    self.contexts = self.get_data_from_setable(self.TableName_Entry.get())
                except Exception as e:
                    showinfo(title = "提示",
                                message = "Table框请输入Seatable TableName！")
                '''
                for i in row_index:
                    try:
                        #context = self.contexts[i]
                        context = self.get_data_from_excel(self.TableName_Entry.get(),i)
                    except Exception as e:
                        showinfo(title = "提示",
                            message = "无法获取数据源，请检查Table框和Excel文件！")
                        raise e
                    try:
                        renames = str(self.File_Name_Entry.get())
                        renames =   renames.replace(' ', '').split('-')
                        Names = ''
                        for rename in renames:             
                            if '{{' in rename and '}}' in rename :
                                rename = rename.replace("{{",'').replace("}}",'')
                                if rename in context.keys():
                                    try:
                                        rename = context[rename]
                                    except Exception as e:
                                        showinfo(title = "提示",
                                            message = "Name框请确定变量名是否正确")
                                else:
                                    showinfo(title = "提示",
                                            message = "命名字段不在Table中")
                            Names = Names + rename
                    except Exception as e:
                        showinfo(title = "提示",
                                    message = "Name框请输入文本")
                        raise e
                    try:
                        self.produce_pdf(context, file_, path_, Names)    
                    except Exception as e:
                        showinfo(title = "提示",
                                message = "请确认模板文件和保存路径已经选择！")
                        raise e       
                showinfo(title = "提示",
                          message = "文件已经生成！请在{0}查看".format(path_))
                os.startfile(path_)
        pass  
    ## 日志函数
    def get_current_time(self):
        current_time = time.strftime('%Y-%m-%d %H:%M:%S',time.localtime(time.time()))
        return current_time
    def write_log_to_Text(self,logmsg):
        global LOG_LINE_NUM
        current_time = self.get_current_time()
        logmsg_in = str(current_time) +" " + str(logmsg) + "\n"      #换行
        if LOG_LINE_NUM <= 9:
            self.log_data_Text.insert(END, logmsg_in)
            LOG_LINE_NUM = LOG_LINE_NUM + 1
        else:
            self.log_data_Text.delete(1.0,2.0)
            self.log_data_Text.insert(END, logmsg_in)
    
def gui_start():
    init_window = Tk()              #实例化出一个父窗口
    ZMJ_PORTAL = MY_GUI(init_window)
    # 设置根窗口默认属性
    ZMJ_PORTAL.set_init_window()

    init_window.mainloop()          #父窗口进入事件循环，可以理解为保持窗口运行，否则界面不展示

if __name__ == "__main__":
    gui_start()